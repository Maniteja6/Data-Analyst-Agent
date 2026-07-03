"""ExcelExporter — exports InsightReport to a multi-sheet XLSX workbook.

Real-time design:
    Excel export runs in a thread pool executor so the asyncio event loop
    is never blocked during workbook construction. Progress is emitted as
    ``report:sheet_complete`` Socket.IO events as each sheet is finalised,
    giving users feedback during large exports.

Workbook structure (5 sheets):
    Executive Summary — KPIs + summary text, brand colours
    Insights          — 5 ranked insights with business impact and confidence
    Data Quality      — completeness, duplicates, null rates per column
    Anomalies         — detected anomalies with row index and description
    Forecast          — forecast table when time series data exists
"""
from __future__ import annotations

import asyncio
import io
from datetime import datetime
from typing import Any

import structlog

logger = structlog.get_logger(__name__)

# DataPilot brand colours (openpyxl ARGB format)
BRAND_VIOLET = "FF5B4FE8"
BRAND_NAVY   = "FF1A1A3E"
WHITE        = "FFFFFFFF"
LIGHT_GREY   = "FFF5F4F1"
DARK_TEXT    = "FF1E1E2E"


async def export_to_excel(
    report: dict,
    sio=None,
    dataset_id: str = "",
) -> bytes:
    """Generate an XLSX workbook from an InsightReport dict.

    Args:
        report:     InsightReport.to_dict() output.
        sio:        Socket.IO server for progress events.
        dataset_id: Dataset UUID for room targeting.

    Returns:
        Raw XLSX bytes ready for S3 upload or streaming download.
    """
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        None, _build_workbook, report, sio, dataset_id, loop
    )


def _build_workbook(
    report: dict,
    sio,
    dataset_id: str,
    loop,
) -> bytes:
    """Synchronous workbook construction (runs in thread pool)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment, Border, Fill, Font, GradientFill,
            PatternFill, Side,
        )
    except ImportError:
        return b"[XLSX export requires openpyxl: pip install openpyxl]"

    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    sheets_done = [0]

    def _emit_progress(sheet_name: str) -> None:
        sheets_done[0] += 1
        if sio and dataset_id:
            try:
                asyncio.run_coroutine_threadsafe(
                    sio.emit(
                        "report:sheet_complete",
                        {"dataset_id": dataset_id, "sheet": sheet_name, "done": sheets_done[0]},
                        room=f"dataset:{dataset_id}",
                    ),
                    loop,
                )
            except Exception:
                pass

    def _header_fill(ws, row, values: list, bg: str = BRAND_VIOLET):
        from openpyxl.styles import Font, PatternFill, Alignment
        fill = PatternFill(fill_type="solid", fgColor=bg)
        font = Font(color=WHITE, bold=True, size=11)
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.fill = fill
            cell.font = font
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: Executive Summary ────────────────────────────────────────
    ws1 = wb.create_sheet("Executive Summary")
    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 50

    ws1["A1"] = "DataPilot Analysis Report"
    ws1["A1"].font = Font(bold=True, size=16, color=BRAND_NAVY.lstrip("FF"))
    ws1["B1"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"

    ws1["A3"] = "Executive Summary"
    ws1["A3"].font = Font(bold=True, size=12)
    ws1.merge_cells("A4:F10")
    ws1["A4"] = report.get("executive_summary", "")
    ws1["A4"].alignment = Alignment(wrap_text=True, vertical="top")

    ws1["A12"] = "Key Metrics"
    ws1["A12"].font = Font(bold=True, size=12)
    _header_fill(ws1, 13, ["Metric", "Value"])

    for i, kpi in enumerate(report.get("kpis", [])[:10], start=14):
        ws1.cell(row=i, column=1, value=kpi.get("name", ""))
        ws1.cell(row=i, column=2, value=kpi.get("value", ""))

    _emit_progress("Executive Summary")

    # ── Sheet 2: Insights ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("Insights")
    _header_fill(ws2, 1, ["#", "Headline", "Business Impact", "Confidence", "Source Columns"])
    ws2.column_dimensions["B"].width = 60
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 15
    ws2.column_dimensions["E"].width = 30

    for i, ins in enumerate(report.get("insights", []), start=2):
        ws2.cell(row=i, column=1, value=i - 1)
        ws2.cell(row=i, column=2, value=ins.get("headline", "")).alignment = Alignment(wrap_text=True)
        ws2.cell(row=i, column=3, value=ins.get("business_impact", "").upper())
        ws2.cell(row=i, column=4, value=f"{float(ins.get('confidence', 0.8))*100:.0f}%")
        ws2.cell(row=i, column=5, value=", ".join(ins.get("source_columns", [])))

        # Colour code by impact
        impact_fill = {
            "high":   PatternFill(fill_type="solid", fgColor="FFFDE68A"),
            "medium": PatternFill(fill_type="solid", fgColor="FFD1FAE5"),
            "low":    PatternFill(fill_type="solid", fgColor="FFF5F4F1"),
        }
        f = impact_fill.get(ins.get("business_impact", "low").lower())
        if f:
            for col in range(1, 6):
                ws2.cell(row=i, column=col).fill = f

    _emit_progress("Insights")

    # ── Sheet 3: Anomalies ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Anomalies")
    _header_fill(ws3, 1, ["Column", "Type", "Severity", "Row #", "Value", "Description"])
    ws3.column_dimensions["F"].width = 60

    for i, alert in enumerate(report.get("anomaly_alerts", [])[:200], start=2):
        ws3.cell(row=i, column=1, value=alert.get("column", ""))
        ws3.cell(row=i, column=2, value=alert.get("anomaly_type", ""))
        ws3.cell(row=i, column=3, value=alert.get("severity", "").upper())
        ws3.cell(row=i, column=4, value=alert.get("row_index", ""))
        ws3.cell(row=i, column=5, value=alert.get("value", ""))
        ws3.cell(row=i, column=6, value=alert.get("description", "")).alignment = Alignment(wrap_text=True)

    _emit_progress("Anomalies")

    # ── Sheet 4: Recommendations ──────────────────────────────────────────
    ws4 = wb.create_sheet("Recommendations")
    _header_fill(ws4, 1, ["Priority", "Title", "Situation", "Action", "Est. Impact"])
    for col, width in [("A", 12), ("B", 30), ("C", 45), ("D", 50), ("E", 25)]:
        ws4.column_dimensions[col].width = width

    for i, rec in enumerate(report.get("recommendations", []), start=2):
        ws4.cell(row=i, column=1, value=rec.get("priority", "").upper())
        ws4.cell(row=i, column=2, value=rec.get("title", ""))
        ws4.cell(row=i, column=3, value=rec.get("situation", "")).alignment = Alignment(wrap_text=True)
        ws4.cell(row=i, column=4, value=rec.get("action", "")).alignment = Alignment(wrap_text=True)
        impact = rec.get("estimated_impact", {})
        impact_str = (
            f"{impact.get('min_pct', 0):.0f}–{impact.get('max_pct', 0):.0f}%"
            if isinstance(impact, dict)
            else str(impact)
        )
        ws4.cell(row=i, column=5, value=impact_str)

    _emit_progress("Recommendations")

    # ── Sheet 5: Forecasts (conditional) ─────────────────────────────────
    forecasts = report.get("forecasts", [])
    if forecasts:
        ws5 = wb.create_sheet("Forecast")
        _header_fill(ws5, 1, ["Date", "Forecast", "Lower Bound", "Upper Bound"])

        row = 2
        for fc in forecasts[:3]:
            for pred in fc.get("predictions", [])[:365]:
                ws5.cell(row=row, column=1, value=pred.get("timestamp", ""))
                ws5.cell(row=row, column=2, value=pred.get("value", 0))
                ws5.cell(row=row, column=3, value=pred.get("lower_bound", 0))
                ws5.cell(row=row, column=4, value=pred.get("upper_bound", 0))
                row += 1

        _emit_progress("Forecast")

    # Serialise to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
